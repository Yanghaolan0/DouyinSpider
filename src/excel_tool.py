import openpyxl
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import time
from pathlib import Path


def _sanitize_for_spreadsheet(value):
    """Neutralize potential spreadsheet formulas when exporting untrusted text."""
    if isinstance(value, str) and value and value[0] in ('=', '+', '-', '@'):
        return "'" + value
    return value


def _sanitize_dataframe_for_spreadsheet(df):
    """Sanitize object/string columns to avoid formula execution in spreadsheet apps."""
    text_cols = df.select_dtypes(include=['object', 'string']).columns
    for col in text_cols:
        df[col] = df[col].map(_sanitize_for_spreadsheet)
    return df

class ExcelTool:
    @staticmethod
    def read_excel_to_list(file_path, sheet_name=None):
        """
        从Excel文件中读取数据到列表
        :param file_path: Excel文件路径
        :param sheet_name: 工作表名称，默认为第一个工作表
        :return: 二维列表，包含Excel中的数据
        """
        data = []
        try:
            workbook = openpyxl.load_workbook(file_path)
            if sheet_name is None:
                sheet = workbook.active
            else:
                sheet = workbook[sheet_name]

            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(list(row))
        except Exception as e:
            print(f"读取Excel文件{file_path}出错: {e}")
        return data

    @staticmethod
    def save_dict_list_to_excel(data, filename):
        """
        将字典数组保存到 Excel 文件中。
        :param data: 字典数组（列表中的字典）
        :param filename: 保存的 Excel 文件名（包括路径）
        """
        # 将字典数组转换为 DataFrame
        df = pd.DataFrame(data)
        df = _sanitize_dataframe_for_spreadsheet(df)
        df = df.fillna('')
        # 将 DataFrame 保存到 Excel 文件
        df.to_excel(filename, index=False)
        print(f"数据已保存到文件: {filename}")

    @staticmethod
    def save_list_to_excel(data, file_path):
        # 将数据转换为DataFrame
        df = pd.DataFrame(data)
        df = _sanitize_dataframe_for_spreadsheet(df)
        
        # 将DataFrame写入Excel文件
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, header=False)
        
        print(f"Data has been written to '{file_path}'")

    @staticmethod
    def read_excel_to_dict(filename):
        # 读取Excel文件
        df = pd.read_excel(filename)
        df = df.fillna('')
        # 将DataFrame转换为字典数组
        dict_array = df.to_dict(orient='records')
        return dict_array
    
    @staticmethod
    def write_dicts_to_excel(dict_list, file_name, key_column):
        """
        将字典数组写入Excel文件。如果文件已存在，则更新或追加数据。

        参数:
        dict_list (list of dict): 要写入的字典数组。
        file_name (str): 输出的Excel文件名。
        """
        try:
            # 将字典列表转换为DataFrame
            new_df = pd.DataFrame(dict_list)
            new_df = _sanitize_dataframe_for_spreadsheet(new_df)
            ensure_parent_dir_exists(file_name)
            
            # 检查文件是否可写，如果不可写则等待5秒后重试
            while True:
                try:
                    if os.path.exists(file_name):
                        with open(file_name, 'a'):
                            pass
                    break
                except IOError:
                    print(f"{file_name} is not writable, retrying in 5 seconds...")
                    time.sleep(5)
            
            # 如果文件已存在，先读取现有数据
            if os.path.exists(file_name):
                existing_df = pd.read_excel(file_name)
                existing_df = _sanitize_dataframe_for_spreadsheet(existing_df)
                # 合并数据，根据'id'列判断是否重复
                combined_df = pd.concat([existing_df, new_df]).drop_duplicates(subset=key_column, keep='last')
            else:
                combined_df = new_df

            combined_df = _sanitize_dataframe_for_spreadsheet(combined_df)
            
            # 将合并后的DataFrame保存到Excel文件
            combined_df.to_excel(file_name, index=False)
            
            # 加载刚刚保存的Excel文件
            workbook = load_workbook(file_name)
            sheet = workbook.active
            
            # 遍历每一列，设置列宽
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter  # 获取列字母
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width
            
            # 保存修改后的Excel文件
            workbook.save(file_name)
            print(f"Data has been written to '{file_name}'")
        except Exception as e:
            print(f"存储excel失败 {e}")


    @staticmethod
    def split_excel_by_sheets(file_path):
        # 加载Excel文件
        xls = pd.ExcelFile(file_path)
        
        # 遍历所有的sheet
        for sheet_name in xls.sheet_names:
            # 读取当前sheet的数据
            df = pd.read_excel(xls, sheet_name=sheet_name)
            # 创建一个新的Excel文件，文件名为sheet名
            new_file_path = f"{sheet_name}.xlsx"
            # 将数据写入新的Excel文件
            with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name=sheet_name)
            print(f"Sheet '{sheet_name}' has been written to '{new_file_path}'")


@staticmethod
def ensure_parent_dir_exists(file_path):
    # 将传入的路径转换为 Path 对象
    path = Path(file_path)

    # 获取父文件夹路径
    parent_dir = path.parent

    # 检查父文件夹是否存在
    if not parent_dir.exists():
        # 如果不存在，则创建父文件夹
        parent_dir.mkdir(parents=True, exist_ok=True)
        print(f"Created directory: {parent_dir}")



# 示例用法
if __name__ == "__main__":
    # 读取Excel文件到列表
    file_path = 'example.xlsx'
    data = ExcelTool.read_excel_to_list(file_path)
    print("读取的数据:", data)

    # 写入列表数据到Excel文件
    new_data = [
        ['Name', 'Age', 'City'],
        ['Alice', 30, 'New York'],
        ['Bob', 25, 'Los Angeles'],
        ['Charlie', 35, 'Chicago']
    ]
    output_file_path = 'output.xlsx'
    ExcelTool.write_list_to_excel(new_data, output_file_path)
    print(f"数据已写入到 {output_file_path}")